"""
CS Templates .xlsx -> data/manual_qa.jsonl
Yüksek kaliteli CS team template'lerini Q&A çiftlerine dönüştürür.
"""
import json
import openpyxl
from pathlib import Path

OUTPUT = Path("data/manual_qa.jsonl")

# Category name -> implied customer question
CATEGORY_QUESTIONS = {
    "Opening":                              "Bonjour",
    "On Hold":                              "Vous pouvez vérifier mon compte ?",
    "Resolution ok":                        "Ma demande a été traitée ?",
    "Other request":                        "Avez-vous autre chose à me dire ?",
    "Closing":                              "Merci, bonne journée",
    "Disconnect from chat (after 2 minutes)": "...",
    "General Information":                  "Quels documents dois-je fournir pour retirer ?",
    "Document sent - In review ":           "J'ai envoyé mes documents, combien de temps pour la vérification ?",
    "How to upload documents":              "Comment envoyer mes documents ?",
    "Problem to add documents from Profile": "Je n'arrive pas à ajouter mes documents depuis mon profil",
    "Missing Documents":                    "Quels documents vous manque-t-il ?",
    "Confirmation delay":                   "Combien de temps pour la vérification des documents ?",
    "Banking card information":             "Quelles informations de carte bancaire faut-il fournir ?",
    "Change request of the mail":           "Je veux changer mon adresse email",
    "Change of address":                    "Je veux changer mon adresse",
    "Change of Phone number":               "Je veux changer mon numéro de téléphone",
    "Change of First Name &/or Last Name":  "Je veux changer mon nom sur le compte",
    "3rd party payment method ":            "Je dois vérifier un mode de paiement tiers",
    "Potential fraud or account migration": "Mon compte a été signalé pour vérification",
    "Delayed on a deposit":                 "Mon dépôt n'est pas arrivé",
    "Proof of deposit ":                    "On me demande une preuve de dépôt",
    "Fees on Deposit":                      "Pourquoi y a-t-il des frais sur mon dépôt ?",
    "Delayed on a withdrawal":              "Mon retrait est en attente depuis longtemps",
    "Withdrawal cancellation":              "Mon retrait a été annulé",
    "Withdrawal limit":                     "Quelles sont les limites de retrait ?",
    "MCO":                                  "Pourquoi je ne peux pas retirer tout mon argent ?",
    "Explanation of MCO":                   "Qu'est-ce que le max cashout ?",
    "Bonus Status/ Check":                  "J'ai reçu mon bonus ?",
    "Bonus Withdrawal / Wager conditions":  "Quelles sont les conditions pour retirer avec un bonus ?",
    "How to activate a bonus":              "Comment activer un bonus ?",
    "Missing Wins":                         "J'ai gagné mais mes gains n'apparaissent pas",
    "Bonus abuser ":                        "Mon bonus a été annulé",
    "Games not counted toward the wager":   "Certains jeux ne comptent pas pour le wagering ?",
    "How can I become VIP?":                "Comment devenir VIP ?",
    "VIP requests being already VIP on\nanother casino": "Je suis déjà VIP sur un autre casino",
    "VIP Online":                           "J'ai une question pour mon manager VIP",
    "VIP Offline":                          "Mon manager VIP n'est pas disponible",
    "Account Closure":                      "Je veux fermer mon compte",
    "Account Reopening / Self Exclusion":   "Je veux réouvrir mon compte",
    "Account Reopening - RED":              "Je veux réouvrir mon compte (addiction)",
    "Following a fraud/locked for security reasons": "Mon compte est bloqué pour des raisons de sécurité",
    "Wrong password lock":                  "Mon compte est bloqué à cause d'un mauvais mot de passe",
    "Technical issue":                      "J'ai un problème technique",
    "Game Error":                           "J'ai une erreur sur un jeu",
    "Legal Threats":                        "Je vais porter plainte",
    "Helplines and support services\n\nGamCare - SOS Joueurs ": "Je veux m'auto-exclure définitivement",
    "My account was hacked":                "Mon compte a été piraté",
    "Tax declaration to authorities":       "Dois-je déclarer mes gains aux impôts ?",
    "Pragmatic games access":               "Je n'ai pas accès aux jeux Pragmatic",
    "Troubleshooting pour le Iphone /Mac ": "Les jeux ne se chargent pas sur iPhone/Mac",
    " All questiosn related to company/license/operator/payment agent": "Qui opère ce casino ?",
}

DEFINITION_QUESTIONS = {
    "Turnover":         "Qu'est-ce que le turnover ?",
    "Wager":            "Comment fonctionne le wagering ?",
    "Max Cash out":     "Qu'est-ce que le max cashout ?",
    "Cashback (VIP)":   "Comment est calculé le cashback VIP ?",
    "Loyalty bonus (VIP)": "Qu'est-ce qu'un loyalty bonus ?",
}

SECTION_HEADERS = {
    "General", "KYC", "BONUSES", "VIP ", "PLAYER ACCOUNTS",
    "LEGAL", "Payments (Deposits/ Withdrawals)", "Customer Requests",
}


def extract_pairs():
    wb = openpyxl.load_workbook("CS Templates .xlsx")
    pairs = []

    # ── Overview sheet ────────────────────────────────────────
    ws = wb["Overview"]
    current_cat = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cat, var, en, fr = row[0], row[1], row[2], row[3]
        if cat:
            current_cat = cat
        if current_cat in SECTION_HEADERS:
            continue

        answer = str(fr).strip() if fr else (str(en).strip() if en else "")
        if not answer or len(answer) < 40:
            continue

        question = CATEGORY_QUESTIONS.get(current_cat)
        if not question:
            continue

        # Remove placeholder names like {name}
        answer = answer.replace("{name}", "notre agent").replace("{brand}", "RomusCasino").replace("{BRAND}", "RomusCasino")

        pairs.append({
            "question": question,
            "answer": answer,
            "category": current_cat,
            "source": "template",
        })

    # ── Definitions sheet ─────────────────────────────────────
    ws2 = wb["Definitions"]
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
        term, explanation = row[0], row[1]
        if not term or not explanation or len(str(explanation)) < 40:
            continue
        question = DEFINITION_QUESTIONS.get(str(term).strip())
        if not question:
            continue
        pairs.append({
            "question": question,
            "answer": str(explanation).strip(),
            "category": "definitions",
            "source": "definition",
        })

    return pairs


def main():
    pairs = extract_pairs()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        for p in pairs:
            f.write(json.dumps(p, ensure_ascii=False) + "\n")
    print(f"Yazildi: {OUTPUT} ({len(pairs)} çift)")
    # Category breakdown
    from collections import Counter
    cats = Counter(p["category"] for p in pairs)
    for cat, n in cats.most_common(10):
        print(f"  {n}x {cat}")


if __name__ == "__main__":
    main()
